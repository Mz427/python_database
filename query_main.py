import os
from openpyxl import load_workbook
import sqlanydb
#import psycopg2
import openpyxl
from datetime import date, timedelta

#Open excel.
today_suffix = date.today().strftime("%Y%m%d")
yesterday = date.today() - timedelta(days=1)
yesterday_suffix = yesterday.strftime("%Y%m%d")
os.chdir("D:/深圳运维工作记录/深圳运维工作记录" + today_suffix + "/zhang/")
excel_fk = load_workbook("发卡网点异常流水监控" + yesterday_suffix + ".xlsx")
sheet_fk = excel_fk.create_sheet(today_suffix[6:8], -2)
cur_sheet_row = 1
cur_sheet_column = 1

#Connect query01.
query_connect = sqlanydb.connect(dsn="query1")
query_cursor = query_connect.cursor()
#Connect postgresql.
#postgresql_connect = psycopg2.connect(database="mzdb", user="mz", password="058474", \
#    host="10.9.7.201" , port="5432")
#postgresql_cursor = postgresql_connect.cursor()

query_sql = " \
    select b.nodename, b.nodeno, a.cardno, a.compno, a.opetime, a.amount, a.balance, a.tracode, \
        a.ctc, a.ttc, a.terminalno, a.wrongflag, b.sinopec_nodeno \
    from addvouch as a, nodeinfor as b  \
    where a.nodeno = b.nodeno and a.wrongflag = 'X' and \
        convert(numeric(8, 0), convert(char(8), a.opetime,112)) = \
        convert(numeric(8, 0), convert(char(8), dateadd(day, -2, getdate()), 112))"
#postgresql_sql = " \
#    select * \
#    from mz_test.mz_addvouch as a\
#    where exitsts (select * from mz_addvouch where a.wrongflag = '1')"

query_cursor.execute(query_sql)
addvouch_x = query_cursor.fetchall()
#postgresql_cursor.execute(postgresql_sql)
#print(postgresql_cursor.fetchall())

for i in addvouch_x :
    #根据后续圈存流水判断是否冲正。
    query_sql = f" \
        select a1.ctc \
        from addvouch as a1 \
        where a1.ctc = {i[8]} and a1.tracode = '{i[7]}' and a1.cardno = '{i[2]}' and a1.wrongflag = '0' \
        union \
        select c1.ctc \
        from carddetail as c1 \
        where c1.ctc = {i[8]} and c1.tracode = '{i[7]}' and c1.cardno = '{i[2]}' and c1.wrongflag = '0'" 
    query_cursor.execute(query_sql)
    temp_result = query_cursor.fetchone()
    if temp_result:
        #插入excel处理结果：写卡失败，待冲正。
        for j in i:
            sheet_fk.cell(row=cur_sheet_row, column=cur_sheet_column).value = j
            cur_sheet_column += 1
        sheet_fk.cell(row=cur_sheet_row, column=cur_sheet_column).value = "写卡失败"
        sheet_fk.cell(row=cur_sheet_row, column=cur_sheet_column+1).value = "待冲正"
        cur_sheet_row += 1
        cur_sheet_column = 1
        continue
    #根据后续圈存流水判断是否确认。
    query_sql = f" \
        select a1.ctc \
        from addvouch as a1 \
        where a1.ctc = {i[8]} + 1 and a1.tracode = '{i[7]}' and a1.cardno = '{i[2]}' and a1.wrongflag = '0' \
        union \
        select c1.ctc \
        from carddetail as c1 \
        where c1.ctc = {i[8]} + 1 and c1.tracode = '{i[7]}' and c1.cardno = '{i[2]}' and c1.wrongflag = '0'" 
    query_cursor.execute(query_sql)
    temp_result = query_cursor.fetchone()
    if temp_result:
        #插入excel处理结果：写卡成功，待确认。
        for j in i:
            sheet_fk.cell(row=cur_sheet_row, column=cur_sheet_column).value = j
            cur_sheet_column += 1
        sheet_fk.cell(row=cur_sheet_row, column=cur_sheet_column).value = "写卡成功"
        sheet_fk.cell(row=cur_sheet_row, column=cur_sheet_column+1).value = "待确认"
        cur_sheet_row += 1
        cur_sheet_column = 1
        continue
    #根据后续加油流水判断是否冲正确认。
    query_sql = f" \
        select * \
        from \
        ( \
            select a3.ctc, a3.balance \
            from oildetail as a3 \
            where a3.cardno = '{i[2]}' and a3.opetime > '{i[4]}' and a3.suctag in ('00', '01', '02') \
            union \
            select b3.ctc, b3.balance \
            from oilvouch as b3 \
            where b3.cardno = '{i[2]}' and b3.opetime > '{i[4]}' and b3.suctag in ('00', '01', '02') \
            union \
            select c3.ctc, c3.balance \
            from unlocal_credit_vouch as c3 \
            where c3.cardno = '{i[2]}' and c3.opetime > '{i[4]}' and c3.suctag in ('00', '01', '02') \
        ) as d3 \
        order by ctc desc" 
    query_cursor.execute(query_sql)
    temp_result = query_cursor.fetchone()
    if temp_result:
        #确定最大CTC卡钱包与卡账是否相等。
        query_sql = f" \
            select a.balance \
            from cardaccount a \
            where a.cardno = '{i[2]}'"
        query_cursor.execute(query_sql)
        temp_result_cardaccount = query_cursor.fetchone()
        if float(temp_result[1]) == float(temp_result_cardaccount[0]):
            #插入excel处理结果：写卡成功，待确认。
            for j in i:
                sheet_fk.cell(row=cur_sheet_row, column=cur_sheet_column).value = j
                cur_sheet_column += 1
            sheet_fk.cell(row=cur_sheet_row, column=cur_sheet_column).value = "写卡成功"
            sheet_fk.cell(row=cur_sheet_row, column=cur_sheet_column+1).value = "待确认"
            cur_sheet_row += 1
            cur_sheet_column = 1
            continue
        elif float(temp_result[1]) == float(temp_result_cardaccount[0][0]) - float(i[5]): 
            #插入excel处理结果：写卡失败，待冲正。
            for j in i:
                sheet_fk.cell(row=cur_sheet_row, column=cur_sheet_column).value = j
                cur_sheet_column += 1
            sheet_fk.cell(row=cur_sheet_row, column=cur_sheet_column).value = "写卡失败"
            sheet_fk.cell(row=cur_sheet_row, column=cur_sheet_column+1).value = "待冲正"
            cur_sheet_row += 1
            cur_sheet_column = 1
            continue
    for j in i:
        sheet_fk.cell(row=cur_sheet_row, column=cur_sheet_column).value = j
        cur_sheet_column += 1
    sheet_fk.cell(row=cur_sheet_row, column=cur_sheet_column).value = "待确认卡钱包"
    sheet_fk.cell(row=cur_sheet_row, column=cur_sheet_column+1).value = "待跟踪"
    cur_sheet_row += 1
    cur_sheet_column = 1

query_cursor.close()
query_connect.close()
excel_fk.save("发卡网点异常流水监控" + today_suffix + ".xlsx")
#postgresql_cursor.close()
#postgresql_connect.close()